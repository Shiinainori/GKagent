#!/usr/bin/env python3
"""生成浙江省普通类高考志愿方案 Excel 文件"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except ImportError:
    print(json.dumps({"error": "缺少 openpyxl，请运行 pip install openpyxl"}, ensure_ascii=False))
    sys.exit(1)


LABEL_COLORS = {
    "rush": "FF6B6B",    # 红 — 冲
    "stable": "4ECDC4",  # 青 — 稳
    "safe": "45B7D1",    # 蓝 — 保
}
LABEL_NAMES = {"rush": "冲", "stable": "稳", "safe": "保"}

HEADERS = [
    "序号", "冲稳保", "学校代号", "学校名称", "专业代号", "专业名称",
    "选考科目要求", "计划招生数", "往年分数", "往年位次", "录取概率",
]

HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def estimate_probability(label: str) -> str:
    return {"rush": "10%-30%", "stable": "40%-60%", "safe": "70%-90%"}.get(label, "未知")


def build_excel(score, rank, subjects, rush_data, stable_data, safe_data, output_dir):
    wb = Workbook()

    # --- 概览 Sheet ---
    ws_summary = wb.active
    ws_summary.title = "方案概览"
    ws_summary.append(["浙江省普通类高考志愿填报方案"])
    ws_summary.append([f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_summary.append([])
    ws_summary.append(["分数", score, "位次", rank, "选考科目", subjects])
    ws_summary.append([])
    ws_summary.append(["类别", "志愿数量"])
    ws_summary.append(["冲", len(rush_data)])
    ws_summary.append(["稳", len(stable_data)])
    ws_summary.append(["保", len(safe_data)])
    ws_summary.append(["合计", len(rush_data) + len(stable_data) + len(safe_data)])
    ws_summary["A1"].font = Font(bold=True, size=16)

    # --- 志愿明细 Sheet ---
    ws_detail = wb.create_sheet("志愿明细")

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    all_data = []
    for label, items in [("rush", rush_data), ("stable", stable_data), ("safe", safe_data)]:
        for item in items:
            all_data.append((label, item))

    for row_idx, (label, item) in enumerate(all_data, 2):
        label_name = LABEL_NAMES.get(label, label)
        prob = estimate_probability(label)
        row_values = [
            row_idx - 1,
            label_name,
            item.get("school_code", ""),
            item.get("school_name", ""),
            item.get("major_code", ""),
            item.get("major_name", ""),
            item.get("subject_requirement", ""),
            item.get("plan_count", ""),
            item.get("score", ""),
            item.get("rank", ""),
            prob,
        ]
        for col_idx, val in enumerate(row_values, 1):
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        color = LABEL_COLORS.get(label)
        if color:
            ws_detail.cell(row=row_idx, column=2).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            ws_detail.cell(row=row_idx, column=2).font = Font(bold=True, color="FFFFFF")

    for col in ws_detail.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_detail.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"GKagent_浙江志愿方案_{score}分.xlsx"
    filepath = output_dir / filename
    wb.save(str(filepath))

    return str(filepath.resolve())


def main():
    parser = argparse.ArgumentParser(description="生成浙江志愿方案 Excel")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--score", type=int, required=True)
    parser.add_argument("--rank", type=int, required=True)
    parser.add_argument("--subjects", required=True)
    parser.add_argument("--rush-data", default="[]")
    parser.add_argument("--stable-data", default="[]")
    parser.add_argument("--safe-data", default="[]")
    args = parser.parse_args()

    rush = json.loads(args.rush_data)
    stable = json.loads(args.stable_data)
    safe = json.loads(args.safe_data)

    filepath = build_excel(
        args.score, args.rank, args.subjects,
        rush, stable, safe, args.output_dir,
    )
    print(json.dumps({"filepath": filepath, "filename": Path(filepath).name}, ensure_ascii=False))


if __name__ == "__main__":
    main()
